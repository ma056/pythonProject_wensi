import openpyxl

# data = pd.read_excel(r'C:\Users\76914\Desktop\data\task\Tesla_Tracking_20221111.xlsm')

# data = openpyxl.load_workbook(filename=r"C:\Users\76914\Desktop\data\task\Tesla_Tracking_20221111.xlsm",
#                               keep_vba=True,
#                               read_only=False)
# sh = data['Tracking']
wb = openpyxl.load_workbook(r"C:\Users\Mawenjing\Desktop\data\task\Tesla_Tracking_20221111.xlsm")
ws = wb['Tracking']
# cell_value_1 = ws.cell(column=1, row=1).value
# set_value_1 = ws.cell(column=1, row=9).value = 866
set_value_2 = ws.cell(column=12, row=11092).value = 6
# 提取区域中的所有单元格对象
# a = 5  # ws.max_column
# cell_3 = ws[f"A2:B{a}"]
#
# # print('A1:C5 的数据对象是', cell_3)
# #
# # for row in cell_3:  # 遍历每一行的单元格
# #     for column in row:  # 遍历每一列的单元格
# #         print(f'{column.value}-----{row}--{column}')  # 提取当前单元格的数据
# for r in range(2, ws.max_column+1):
#     data_a = ws.cell(row=r, column=5).value
#     data_b = ws.cell(row=r, column=6).value
#     project_name_locale = f"{data_a}_'{data_b}"
# update_dict = {}
# for r in range(2, 5):
#     data_a = ws.cell(row=r, column=5).value
#     data_b = ws.cell(row=r, column=6).value
#     project_name_locale = f"{data_a}_'{data_b}"
#     update_dict[project_name_locale] = r
# print(update_dict)
# a = {'a': [1, 2, 3], 'b': [4, 1, 4]}
# for key, value1 in a.items():
#     r = value1[-1]
#     for clo in range(7, 9):
#         # print(clo,r,value1[clo-7])
#         set_value_2 = ws.cell(column=clo, row=r).value = value1[clo-7]

# print('cell_value_1:',cell_value_1)
wb.save('cell_operation1.xlsm')
wb.close()
