import pandas as pd
import os
import openpyxl


def main(holiday_paths, task_paths):
    info_dict = {}
    for file_name in os.listdir(holiday_paths):
        project_name = '_'.join(file_name.split('_')[:4])
        df = pd.read_csv(os.path.join(holiday_paths, file_name))
        target_locale = df['Target Locale'][0]
        data = df.loc[1, "Total":"Multiple 100%"].values.tolist()
        data1 = df['Cost Estimate (USD)'].values.tolist()[2:]
        data.extend(data1)
        key = f'{project_name}_{target_locale}'
        info_dict[key] = data
    # print(info_dict)
    wb = openpyxl.load_workbook(task_paths)
    ws = wb['Tracking']
    for r in range(2, ws.max_column):
        data_a = ws.cell(row=r, column=5).value
        data_b = ws.cell(row=r, column=6).value
        project_name_locale = f"{data_a}_'{data_b}"
        if project_name_locale in info_dict:
            info_dict[project_name_locale].append(r)
    # # l - z     12 - 26
    # print(info_dict)
    for key, value1 in info_dict.items():
        r = value1[-1]
        for clo in range(12, 27):
            # print(clo,r,clo - 12)
            ws.cell(column=clo, row=r).value = value1[clo - 12]
    wb.save(os.path.basename(task_paths))
    wb.close()


if __name__ == '__main__':
    task_paths = r"C:\Users\Mawenjing\Desktop\data\task\Tesla_Tracking_20221111.xlsm"
    holiday_paths = r'C:\Users\Mawenjing\Desktop\data\11.11 holiday'
    main(holiday_paths, task_paths)
